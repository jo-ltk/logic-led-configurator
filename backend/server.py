from dotenv import load_dotenv
from pathlib import Path
load_dotenv(Path(__file__).parent / ".env")

import os
import uuid
import logging
from datetime import datetime, timezone
from fastapi import FastAPI, APIRouter, HTTPException, Response, Request, Depends, Query
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from io import BytesIO
from typing import Optional, List

from db import db
from auth_utils import (
    hash_password, verify_password, create_access_token, create_refresh_token,
    decode_token, get_current_user, require_roles,
)
from models import (
    UserCreate, UserLogin, Product, Customer, Partner, Quote, QuoteCreate,
    CompanySettings,
)
from calculator import to_mm, calc_wall, build_boq
from pdf_gen import generate_quote_pdf
from seed import seed_all

app = FastAPI(title="LOGIC LED Videowall Configurator API")
api = APIRouter(prefix="/api")

logger = logging.getLogger("uvicorn.error")


@app.on_event("startup")
async def on_start():
    try:
        await seed_all()
        logger.info("Seed complete")
    except Exception as e:
        logger.error(f"Seed skipped: {e}")
    try:
        from storage import init_storage
        init_storage()
    except Exception as e:
        logger.error(f"Storage skipped: {e}")


@app.get("/health")
async def health():
    return {"status": "ok"}


@api.get("/health")
async def api_health():
    return {"status": "ok"}


@api.get("/")
async def root():
    return {"message": "LOGIC LED API", "version": "1.0"}


def _set_cookies(resp: Response, access: str, refresh: str):
    resp.set_cookie("access_token", access, httponly=True, secure=False, samesite="lax",
                    max_age=8 * 3600, path="/")
    resp.set_cookie("refresh_token", refresh, httponly=True, secure=False, samesite="lax",
                    max_age=7 * 86400, path="/")


@api.post("/auth/login")
async def login(body: UserLogin, response: Response):
    email = body.email.lower().strip()
    user = await db.users.find_one({"email": email})
    if not user or not verify_password(body.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    access = create_access_token(user["id"], user["email"], user["role"])
    refresh = create_refresh_token(user["id"])
    _set_cookies(response, access, refresh)
    return {"id": user["id"], "email": user["email"], "name": user["name"],
            "role": user["role"], "access_token": access}


@api.post("/auth/logout")
async def logout(response: Response):
    response.delete_cookie("access_token", path="/")
    response.delete_cookie("refresh_token", path="/")
    return {"ok": True}


@api.get("/auth/me")
async def me(user=Depends(get_current_user)):
    return user


@api.get("/auth/users")
async def list_users(role: Optional[str] = None, user=Depends(get_current_user)):
    q: dict = {}
    if role: q["role"] = role
    users = await db.users.find(q, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users


# ---------- PRODUCTS ----------
@api.get("/products")
async def list_products(category: Optional[str] = None, pixel_pitch: Optional[str] = None,
                        include_archived: bool = False, user=Depends(get_current_user)):
    q: dict = {}
    if category: q["category"] = category
    if pixel_pitch: q["pixel_pitch"] = pixel_pitch
    if not include_archived: q["archived"] = False
    return await db.products.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api.get("/products/{pid}")
async def get_product(pid: str, user=Depends(get_current_user)):
    doc = await db.products.find_one({"id": pid}, {"_id": 0})
    if not doc: raise HTTPException(404, "Product not found")
    return doc


@api.post("/products")
async def create_product(body: Product, user=Depends(require_roles("product_manager", "super_admin"))):
    doc = body.model_dump()
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/products/{pid}")
async def update_product(pid: str, body: Product, user=Depends(require_roles("product_manager", "super_admin"))):
    d = body.model_dump(); d["id"] = pid
    await db.products.update_one({"id": pid}, {"$set": d})
    return d


@api.delete("/products/{pid}")
async def delete_product(pid: str, user=Depends(require_roles("product_manager", "super_admin"))):
    await db.products.update_one({"id": pid}, {"$set": {"archived": True}})
    return {"ok": True}


@api.post("/products/{pid}/duplicate")
async def duplicate_product(pid: str, user=Depends(require_roles("product_manager", "super_admin"))):
    doc = await db.products.find_one({"id": pid}, {"_id": 0})
    if not doc: raise HTTPException(404, "Product not found")
    doc["id"] = str(uuid.uuid4())
    doc["name"] = doc["name"] + " (Copy)"
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.products.insert_one(doc)
    doc.pop("_id", None)
    return doc


# ---------- CUSTOMERS ----------
@api.get("/customers")
async def list_customers(user=Depends(get_current_user)):
    rows = await db.customers.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    quotes = await db.quotes.find({}, {"_id": 0, "customer_id": 1, "grand_total": 1, "status": 1}).to_list(5000)
    from collections import defaultdict
    counts = defaultdict(lambda: {"count": 0, "value": 0, "won": 0})
    for q in quotes:
        c = counts[q.get("customer_id")]
        c["count"] += 1
        c["value"] += q.get("grand_total", 0)
        if q.get("status") == "won": c["won"] += 1
    for r in rows:
        r.update(counts.get(r["id"], {"count": 0, "value": 0, "won": 0}))
    return rows


@api.get("/customers/{cid}/quotes")
async def customer_quotes(cid: str, user=Depends(get_current_user)):
    return await db.quotes.find({"customer_id": cid}, {"_id": 0}).sort("created_at", -1).to_list(500)


@api.post("/customers")
async def create_customer(body: Customer, user=Depends(get_current_user)):
    doc = body.model_dump()
    await db.customers.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/customers/{cid}")
async def update_customer(cid: str, body: Customer, user=Depends(get_current_user)):
    d = body.model_dump(); d["id"] = cid
    await db.customers.update_one({"id": cid}, {"$set": d})
    return d


@api.delete("/customers/{cid}")
async def delete_customer(cid: str, user=Depends(require_roles("super_admin", "sales_manager"))):
    await db.customers.delete_one({"id": cid})
    return {"ok": True}


# ---------- PARTNERS ----------
@api.get("/partners")
async def list_partners(user=Depends(get_current_user)):
    rows = await db.partners.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    quotes = await db.quotes.find({}, {"_id": 0, "partner_id": 1, "grand_total": 1, "status": 1}).to_list(5000)
    from collections import defaultdict
    counts = defaultdict(lambda: {"count": 0, "value": 0, "won": 0})
    for q in quotes:
        pid = q.get("partner_id")
        if not pid: continue
        c = counts[pid]
        c["count"] += 1
        c["value"] += q.get("grand_total", 0)
        if q.get("status") == "won": c["won"] += 1
    for r in rows:
        r.update(counts.get(r["id"], {"count": 0, "value": 0, "won": 0}))
    return rows


@api.get("/partners/{pid}/quotes")
async def partner_quotes(pid: str, user=Depends(get_current_user)):
    return await db.quotes.find({"partner_id": pid}, {"_id": 0}).sort("created_at", -1).to_list(500)


@api.post("/partners")
async def create_partner(body: Partner, user=Depends(get_current_user)):
    doc = body.model_dump()
    await db.partners.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.put("/partners/{pid}")
async def update_partner(pid: str, body: Partner, user=Depends(get_current_user)):
    d = body.model_dump(); d["id"] = pid
    await db.partners.update_one({"id": pid}, {"$set": d})
    return d


@api.delete("/partners/{pid}")
async def delete_partner(pid: str, user=Depends(require_roles("super_admin", "sales_manager"))):
    await db.partners.delete_one({"id": pid})
    return {"ok": True}


# ---------- CONFIGURATOR ----------
from pydantic import BaseModel


class ConfigInput(BaseModel):
    product_id: str
    width: float
    height: float
    unit: str = "mm"
    include: Optional[dict] = None
    spare_percent: Optional[float] = 2.0


@api.post("/configure")
async def configure(body: ConfigInput, user=Depends(get_current_user)):
    product = await db.products.find_one({"id": body.product_id}, {"_id": 0})
    if not product:
        raise HTTPException(404, "Product not found")
    wmm = to_mm(body.width, body.unit)
    hmm = to_mm(body.height, body.unit)
    if wmm <= 0 or hmm <= 0:
        raise HTTPException(400, "Invalid dimensions")
    wall = calc_wall(product, wmm, hmm)
    settings = await db.settings.find_one({"id": "default"}, {"_id": 0}) or {}
    items = build_boq(product, wall, settings, body.include, body.spare_percent or 2.0)
    subtotal = sum(i["total"] for i in items)
    gst_pct = settings.get("default_gst_percent", 18)
    gst_amt = round(subtotal * gst_pct / 100, 2)
    grand = round(subtotal + gst_amt, 2)
    return {
        "product": product, "wall": wall, "items": items,
        "subtotal": subtotal, "gst_percent": gst_pct, "gst_amount": gst_amt,
        "grand_total": grand, "width_mm": wmm, "height_mm": hmm,
    }


# ---------- QUOTES ----------
async def _next_quote_number() -> str:
    settings = await db.settings.find_one({"id": "default"}, {"_id": 0}) or {}
    prefix = settings.get("quote_prefix", "LOGIC/Q")
    year = datetime.now().year
    count = await db.quotes.count_documents({}) + 1
    return f"{prefix}/{year}/{count:04d}"


@api.get("/quotes")
async def list_quotes(status: Optional[str] = None, user=Depends(get_current_user)):
    q: dict = {}
    if status: q["status"] = status
    if user["role"] == "sales_executive":
        q["sales_person_id"] = user["id"]
    return await db.quotes.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api.get("/quotes/{qid}")
async def get_quote(qid: str, user=Depends(get_current_user)):
    doc = await db.quotes.find_one({"id": qid}, {"_id": 0})
    if not doc: raise HTTPException(404, "Quote not found")
    return doc


@api.post("/quotes")
async def create_quote(body: QuoteCreate, user=Depends(get_current_user)):
    d = body.model_dump()
    d["id"] = str(uuid.uuid4())
    d["quote_number"] = await _next_quote_number()
    d["revision"] = 0
    d["status"] = "draft"
    # Allow sales-manager / super-admin to attribute quote to another rep
    override_id = getattr(body, "sales_person_id_override", None) or d.pop("sales_person_id_override", None)
    if override_id and user["role"] in ("super_admin", "sales_manager"):
        rep = await db.users.find_one({"id": override_id}, {"_id": 0, "password_hash": 0})
        if rep:
            d["sales_person_id"] = rep["id"]
            d["sales_person_name"] = rep["name"]
        else:
            d["sales_person_id"] = user["id"]; d["sales_person_name"] = user["name"]
    else:
        d["sales_person_id"] = user["id"]
        d["sales_person_name"] = user["name"]
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.quotes.insert_one(d)
    d.pop("_id", None)
    return d


@api.patch("/quotes/{qid}/status")
async def update_quote_status(qid: str, status: str, user=Depends(get_current_user)):
    if status not in ("draft", "submitted", "won", "lost", "cancelled"):
        raise HTTPException(400, "Invalid status")
    await db.quotes.update_one({"id": qid}, {"$set": {"status": status}})
    return {"ok": True, "status": status}


@api.delete("/quotes/{qid}")
async def delete_quote(qid: str, user=Depends(require_roles("super_admin", "sales_manager"))):
    await db.quotes.delete_one({"id": qid})
    return {"ok": True}


@api.get("/quotes/{qid}/pdf")
async def quote_pdf(qid: str, user=Depends(get_current_user)):
    quote = await db.quotes.find_one({"id": qid}, {"_id": 0})
    if not quote: raise HTTPException(404, "Quote not found")
    product = await db.products.find_one({"id": quote["config"]["product_id"]}, {"_id": 0})
    customer = await db.customers.find_one({"id": quote["customer_id"]}, {"_id": 0}) or {}
    settings = await db.settings.find_one({"id": "default"}, {"_id": 0}) or CompanySettings().model_dump()

    # Read enriched saved fields (if quote has attached engineering + render)
    power = quote.get("power_data")
    network = quote.get("network_data")
    controller = quote.get("controller_data")
    sections = quote.get("pdf_sections") or {}
    cover_intro = quote.get("cover_intro")

    # Load render image if attached
    render_image_bytes = None
    if quote.get("render_file_id"):
        from storage import get_object
        try:
            rec = await db.files.find_one({"id": quote["render_file_id"]}, {"_id": 0})
            if rec:
                render_image_bytes, _ = get_object(rec["storage_path"])
        except Exception:
            pass

    # Load product image
    product_image_bytes = None
    if product and product.get("image_file_id"):
        from storage import get_object
        try:
            rec = await db.files.find_one({"id": product["image_file_id"]}, {"_id": 0})
            if rec:
                product_image_bytes, _ = get_object(rec["storage_path"])
        except Exception:
            pass

    pdf_bytes = generate_quote_pdf(quote, product, customer, settings,
                                   power=power, network=network, controller=controller,
                                   render_image_bytes=render_image_bytes,
                                   product_image_bytes=product_image_bytes,
                                   cover_intro=cover_intro, sections=sections)
    fname = quote["quote_number"].replace("/", "_")
    return StreamingResponse(BytesIO(pdf_bytes), media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}.pdf"'})


@api.get("/quotes/{qid}/excel")
async def quote_excel(qid: str, user=Depends(get_current_user)):
    from openpyxl import Workbook
    quote = await db.quotes.find_one({"id": qid}, {"_id": 0})
    if not quote: raise HTTPException(404, "Quote not found")
    product = await db.products.find_one({"id": quote["config"]["product_id"]}, {"_id": 0})
    customer = await db.customers.find_one({"id": quote["customer_id"]}, {"_id": 0}) or {}
    wb = Workbook()
    ws = wb.active; ws.title = "Cover"
    ws["A1"] = "LOGIC LED Quotation"
    ws["A3"] = "Quote #"; ws["B3"] = quote["quote_number"]
    ws["A4"] = "Project"; ws["B4"] = quote["project_name"]
    ws["A5"] = "Customer"; ws["B5"] = customer.get("company", "")
    ws["A6"] = "Date"; ws["B6"] = quote["created_at"][:10]
    ws["A7"] = "Grand Total"; ws["B7"] = quote["grand_total"]
    ws2 = wb.create_sheet("BOQ")
    ws2.append(["#", "Description", "Qty", "Unit", "Rate", "Amount"])
    for i, it in enumerate(quote["items"], 1):
        ws2.append([i, it["description"], it["qty"], it["unit"], it["unit_price"], it["total"]])
    ws2.append(["", "Subtotal", "", "", "", quote["subtotal"]])
    ws2.append(["", f"GST @ {quote['gst_percent']}%", "", "", "", quote["gst_amount"]])
    ws2.append(["", "Grand Total", "", "", "", quote["grand_total"]])
    ws3 = wb.create_sheet("Configuration")
    for k, v in quote["config"].items():
        ws3.append([k, v])
    ws4 = wb.create_sheet("Specifications")
    if product:
        for k, v in product.items():
            ws4.append([k, str(v)])
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    fname = quote["quote_number"].replace("/", "_")
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})


# ---------- DASHBOARD ----------
@api.get("/dashboard/stats")
async def dashboard_stats(user=Depends(get_current_user)):
    q_all = await db.quotes.find({}, {"_id": 0}).to_list(2000)
    customers = await db.customers.count_documents({})
    partners = await db.partners.count_documents({})
    total_value = sum(q["grand_total"] for q in q_all)
    total_area = sum(q["config"]["led_area_sqm"] for q in q_all)
    from collections import Counter, defaultdict
    pitches = Counter(); series = Counter(); categories = Counter()
    products_map = {p["id"]: p for p in await db.products.find({}, {"_id": 0}).to_list(500)}
    for q in q_all:
        p = products_map.get(q["config"]["product_id"])
        if p:
            pitches[p["pixel_pitch"]] += 1
            series[p["series"]] += 1
            categories[p["category"]] += 1
    monthly = defaultdict(lambda: {"count": 0, "value": 0})
    for q in q_all:
        m = q["created_at"][:7]
        monthly[m]["count"] += 1
        monthly[m]["value"] += q["grand_total"]
    monthly_series = [{"month": k, **v} for k, v in sorted(monthly.items())][-6:]
    status_counts = Counter(q["status"] for q in q_all)
    in_out = {"indoor": 0, "outdoor": 0, "other": 0}
    for q in q_all:
        p = products_map.get(q["config"]["product_id"])
        if p:
            if "Indoor" in p["category"]: in_out["indoor"] += 1
            elif "Outdoor" in p["category"]: in_out["outdoor"] += 1
            else: in_out["other"] += 1
    region = Counter(q.get("region") or "N/A" for q in q_all)
    sp = defaultdict(lambda: {"count": 0, "value": 0})
    for q in q_all:
        sp[q["sales_person_name"]]["count"] += 1
        sp[q["sales_person_name"]]["value"] += q["grand_total"]
    salesperson = [{"name": k, **v} for k, v in sp.items()]
    recent = sorted(q_all, key=lambda x: x["created_at"], reverse=True)[:5]
    return {
        "total_quotations": len(q_all),
        "total_customers": customers,
        "total_partners": partners,
        "total_sales_value": total_value,
        "total_led_area": round(total_area, 2),
        "most_sold_pitch": pitches.most_common(1)[0][0] if pitches else "-",
        "most_sold_series": series.most_common(1)[0][0] if series else "-",
        "monthly_trend": monthly_series,
        "status_counts": dict(status_counts),
        "pitch_distribution": [{"pitch": k, "count": v} for k, v in pitches.items()],
        "category_distribution": [{"category": k, "count": v} for k, v in categories.items()],
        "region_distribution": [{"region": k, "count": v} for k, v in region.items()],
        "indoor_outdoor": in_out,
        "salesperson_performance": salesperson,
        "recent_quotes": [{"id": q["id"], "quote_number": q["quote_number"],
                           "project_name": q["project_name"], "grand_total": q["grand_total"],
                           "status": q["status"], "created_at": q["created_at"]} for q in recent],
    }


# ---------- SETTINGS ----------
@api.get("/settings")
async def get_settings(user=Depends(get_current_user)):
    doc = await db.settings.find_one({"id": "default"}, {"_id": 0})
    if not doc:
        doc = CompanySettings().model_dump()
        await db.settings.insert_one(doc)
    return doc


@api.put("/settings")
async def update_settings(body: CompanySettings, user=Depends(require_roles("super_admin"))):
    d = body.model_dump(); d["id"] = "default"
    await db.settings.update_one({"id": "default"}, {"$set": d}, upsert=True)
    return d


# ---------- AI ----------
class AIRecInput(BaseModel):
    application: str
    viewing_distance_m: float
    indoor: bool = True
    budget_range_inr: Optional[str] = None
    ambient_light: Optional[str] = None


@api.post("/ai/recommend")
async def ai_recommend(body: AIRecInput, user=Depends(get_current_user)):
    products = await db.products.find({"archived": False}, {"_id": 0}).to_list(50)
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    key = os.environ["EMERGENT_LLM_KEY"]
    sys_msg = (
        "You are an expert LED videowall consultant for LOGIC LED Displays. "
        "Recommend the best product from the provided catalog for the customer's use case. "
        "Consider viewing distance (optimal ≈ pixel pitch mm × 2.5 in meters), "
        "brightness (indoor 600-1200 nits, outdoor 4500+ nits), and application. "
        'Respond strictly with JSON: {"recommended_product_id": "<id>", "reason": "<max 3 sentences>", "alternatives": ["<id>", "<id>"]}'
    )
    catalog = [{"id": p["id"], "name": p["name"], "category": p["category"],
                "pixel_pitch": p["pixel_pitch"], "brightness_nits": p["brightness_nits"],
                "cabinet_price": p["cabinet_price"]} for p in products]
    prompt = (
        f"Application: {body.application}\nViewing distance: {body.viewing_distance_m} m\n"
        f"Indoor: {body.indoor}\nBudget: {body.budget_range_inr or 'flexible'}\n"
        f"Ambient light: {body.ambient_light or 'normal'}\n\nCatalog:\n{catalog}\n\nReturn only JSON."
    )
    chat = LlmChat(api_key=key, session_id=f"rec-{user['id']}", system_message=sys_msg).with_model("anthropic", "claude-sonnet-4-6")
    try:
        reply = await chat.send_message(UserMessage(text=prompt))
    except Exception as e:
        msg = str(e)
        if "Budget" in msg or "budget" in msg:
            raise HTTPException(status_code=402, detail="AI budget exceeded. Please top up your Emergent LLM key in Profile → Universal Key.")
        raise HTTPException(status_code=503, detail=f"AI service unavailable: {msg[:200]}")
    import json, re
    text = reply if isinstance(reply, str) else str(reply)
    m = re.search(r"\{.*\}", text, re.DOTALL)
    parsed = {}
    if m:
        try: parsed = json.loads(m.group(0))
        except Exception: parsed = {}
    return {"raw": text, "parsed": parsed}


class AISummaryInput(BaseModel):
    quote_id: str


@api.post("/ai/proposal-summary")
async def ai_proposal_summary(body: AISummaryInput, user=Depends(get_current_user)):
    quote = await db.quotes.find_one({"id": body.quote_id}, {"_id": 0})
    if not quote: raise HTTPException(404, "Quote not found")
    product = await db.products.find_one({"id": quote["config"]["product_id"]}, {"_id": 0})
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    key = os.environ["EMERGENT_LLM_KEY"]
    sys_msg = "You are a technical writer for LOGIC LED Displays. Write a professional executive summary for a videowall proposal in 4-6 sentences."
    prompt = (f"Project: {quote['project_name']}\nProduct: {product['name']}\n"
              f"Size: {quote['config']['total_width_mm']/1000:.2f}m × {quote['config']['total_height_mm']/1000:.2f}m\n"
              f"Resolution: {quote['config']['resolution_w']}×{quote['config']['resolution_h']}\n"
              f"Cabinets: {quote['config']['total_cabinets']}\nTotal: INR {quote['grand_total']:,.0f}")
    chat = LlmChat(api_key=key, session_id=f"sum-{body.quote_id}", system_message=sys_msg).with_model("anthropic", "claude-sonnet-4-6")
    try:
        text = await chat.send_message(UserMessage(text=prompt))
    except Exception as e:
        msg = str(e)
        if "Budget" in msg or "budget" in msg:
            raise HTTPException(status_code=402, detail="AI budget exceeded. Please top up your Emergent LLM key.")
        raise HTTPException(status_code=503, detail=f"AI service unavailable: {msg[:200]}")
    return {"summary": text if isinstance(text, str) else str(text)}


# ---------- REPORTS ----------
@api.get("/reports/sales")
async def report_sales(period: str = "monthly", user=Depends(get_current_user)):
    from collections import defaultdict
    q_all = await db.quotes.find({}, {"_id": 0}).to_list(2000)
    buckets = defaultdict(lambda: {"count": 0, "value": 0, "won_value": 0})
    for q in q_all:
        if period == "yearly": key = q["created_at"][:4]
        elif period == "quarterly":
            y, m = q["created_at"][:4], int(q["created_at"][5:7])
            key = f"{y}-Q{(m-1)//3+1}"
        else: key = q["created_at"][:7]
        buckets[key]["count"] += 1
        buckets[key]["value"] += q["grand_total"]
        if q["status"] == "won": buckets[key]["won_value"] += q["grand_total"]
    return [{"period": k, **v} for k, v in sorted(buckets.items())]


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

# ==================== NEW MODULES ====================

from controllers_engine import NOVASTAR_CATALOG, plan_controller
from power_engine import power_calc
from network_engine import network_calc
from structural import structural_calc
from roles import ROLES, ROLE_LABELS, has_permission
from storage import put_object, get_object, save_file_record, APP_NAME

api2 = APIRouter(prefix="/api")


# --- Meta ---
@api2.get("/meta/roles")
async def meta_roles():
    return {"roles": ROLES, "labels": ROLE_LABELS}


# --- User management (admin panel) ---
class UserAdminCreate(BaseModel):
    email: str
    password: str
    name: str
    role: str
    active: bool = True


class UserAdminUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    active: Optional[bool] = None
    password: Optional[str] = None


@api2.post("/admin/users")
async def create_user(body: UserAdminCreate, user=Depends(require_roles("super_admin", "admin"))):
    if body.role not in ROLES: raise HTTPException(400, "Invalid role")
    if await db.users.find_one({"email": body.email.lower()}):
        raise HTTPException(409, "Email already exists")
    doc = {
        "id": str(uuid.uuid4()), "email": body.email.lower(), "name": body.name,
        "role": body.role, "active": body.active,
        "password_hash": hash_password(body.password),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(doc)
    doc.pop("_id", None); doc.pop("password_hash", None)
    return doc


@api2.put("/admin/users/{uid}")
async def update_user(uid: str, body: UserAdminUpdate, user=Depends(require_roles("super_admin", "admin"))):
    upd = {}
    if body.name is not None: upd["name"] = body.name
    if body.role is not None:
        if body.role not in ROLES: raise HTTPException(400, "Invalid role")
        upd["role"] = body.role
    if body.active is not None: upd["active"] = body.active
    if body.password: upd["password_hash"] = hash_password(body.password)
    await db.users.update_one({"id": uid}, {"$set": upd})
    doc = await db.users.find_one({"id": uid}, {"_id": 0, "password_hash": 0})
    return doc


@api2.delete("/admin/users/{uid}")
async def delete_user(uid: str, user=Depends(require_roles("super_admin"))):
    if user["id"] == uid: raise HTTPException(400, "Cannot delete yourself")
    await db.users.delete_one({"id": uid})
    return {"ok": True}


# --- Novastar controllers ---
@api2.get("/controllers")
async def list_controllers(user=Depends(get_current_user)):
    return await db.controllers.find({}, {"_id": 0}).sort("max_pixels", 1).to_list(50)


@api2.get("/controllers/catalog")
async def controllers_catalog(user=Depends(get_current_user)):
    return NOVASTAR_CATALOG


class ControllerCreate(BaseModel):
    model: str
    class_: str = "Custom"
    max_pixels: int
    output_ports: int
    pixels_per_port: int
    features: List[str] = []
    price: float = 0


class ControllerUpdate(BaseModel):
    model: Optional[str] = None
    class_: Optional[str] = None
    max_pixels: Optional[int] = None
    output_ports: Optional[int] = None
    pixels_per_port: Optional[int] = None
    features: Optional[List[str]] = None
    price: Optional[float] = None


@api2.post("/controllers")
async def create_controller(body: ControllerCreate, user=Depends(require_roles("super_admin", "admin"))):
    doc = body.model_dump()
    doc["class"] = doc.pop("class_", "Custom")
    doc["id"] = str(uuid.uuid4())
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    await db.controllers.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api2.put("/controllers/{cid}")
async def update_controller(cid: str, body: ControllerUpdate, user=Depends(require_roles("super_admin", "admin"))):
    upd = {k: v for k, v in body.model_dump().items() if v is not None}
    if "class_" in upd: upd["class"] = upd.pop("class_")
    await db.controllers.update_one({"id": cid}, {"$set": upd})
    doc = await db.controllers.find_one({"id": cid}, {"_id": 0})
    return doc


@api2.delete("/controllers/{cid}")
async def delete_controller(cid: str, user=Depends(require_roles("super_admin"))):
    await db.controllers.delete_one({"id": cid})
    return {"ok": True}


class ControllerPlanInput(BaseModel):
    total_pixels: int
    prefer_class: Optional[str] = None


@api2.post("/controllers/plan")
async def controllers_plan(body: ControllerPlanInput, user=Depends(get_current_user)):
    return plan_controller(body.total_pixels, body.prefer_class)


# --- Power engineering ---
class PowerInput(BaseModel):
    power_max_kw: float
    power_typical_kw: float
    voltage: int = 230
    phase: str = "single"  # single | three
    daily_hours: float = 10
    tariff_inr_kwh: float = 8.5
    days_per_month: int = 30


@api2.post("/engineering/power")
async def engineering_power(body: PowerInput, user=Depends(get_current_user)):
    return power_calc(body.power_max_kw, body.power_typical_kw, body.voltage,
                      body.phase, body.daily_hours, body.tariff_inr_kwh, body.days_per_month)


# --- Network engineering ---
class NetworkInput(BaseModel):
    cabinets_x: int
    cabinets_y: int
    total_pixels: int
    controller_ports: int = 16
    pixels_per_port: int = 650000
    cabinet_pitch_mm: float = 500
    run_length_m: float = 30
    environment: str = "indoor"  # indoor | outdoor
    pixels_per_cabinet: Optional[int] = None


@api2.post("/engineering/network")
async def engineering_network(body: NetworkInput, user=Depends(get_current_user)):
    return network_calc(body.cabinets_x, body.cabinets_y, body.controller_ports,
                        body.pixels_per_port, body.total_pixels,
                        body.cabinet_pitch_mm, body.run_length_m,
                        body.environment, body.pixels_per_cabinet)


# --- Structural engineering ---
class StructuralInput(BaseModel):
    total_weight_kg: float
    wall_w_mm: float
    wall_h_mm: float
    total_cabinets: int
    mount_type: str = "wall_mount"  # wall_mount | hanging | floor_stand


@api2.post("/engineering/structural")
async def engineering_structural(body: StructuralInput, user=Depends(get_current_user)):
    return structural_calc(body.total_weight_kg, body.wall_w_mm, body.wall_h_mm,
                           body.total_cabinets, body.mount_type)


# --- Projects (pipeline) ---
PROJECT_STAGES = ["enquiry", "quotation", "po", "manufacturing", "qc", "dispatch",
                  "installation", "commissioning", "amc", "warranty"]


class ProjectCreate(BaseModel):
    quote_id: Optional[str] = None
    customer_id: str
    name: str
    stage: str = "enquiry"
    owner_id: Optional[str] = None
    notes: Optional[str] = None
    expected_close_date: Optional[str] = None
    value: float = 0


@api2.get("/projects")
async def list_projects(stage: Optional[str] = None, user=Depends(get_current_user)):
    q = {}
    if stage: q["stage"] = stage
    return await db.projects.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)


@api2.post("/projects")
async def create_project(body: ProjectCreate, user=Depends(get_current_user)):
    if body.stage not in PROJECT_STAGES:
        raise HTTPException(400, "Invalid stage")
    d = body.model_dump()
    d["id"] = str(uuid.uuid4())
    d["created_at"] = datetime.now(timezone.utc).isoformat()
    d["created_by"] = user["id"]
    await db.projects.insert_one(d)
    d.pop("_id", None)
    return d


@api2.patch("/projects/{pid}/stage")
async def update_project_stage(pid: str, stage: str, user=Depends(get_current_user)):
    if stage not in PROJECT_STAGES:
        raise HTTPException(400, "Invalid stage")
    await db.projects.update_one({"id": pid}, {"$set": {"stage": stage,
                                                        "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"ok": True, "stage": stage}


@api2.delete("/projects/{pid}")
async def delete_project(pid: str, user=Depends(require_roles("super_admin", "admin"))):
    await db.projects.delete_one({"id": pid})
    return {"ok": True}


@api2.get("/projects/stages")
async def project_stages(user=Depends(get_current_user)):
    return {"stages": PROJECT_STAGES}


# --- File upload ---
from fastapi import UploadFile, File, Header, Query
from fastapi.responses import Response


@api2.post("/upload")
async def upload_file(kind: str = "generic", file: UploadFile = File(...), user=Depends(get_current_user)):
    ext = file.filename.rsplit(".", 1)[-1] if "." in file.filename else "bin"
    path = f"{APP_NAME}/uploads/{user['id']}/{uuid.uuid4()}.{ext}"
    data = await file.read()
    try:
        result = put_object(path, data, file.content_type or "application/octet-stream")
    except Exception as e:
        raise HTTPException(503, f"Storage error: {str(e)[:200]}")
    rec = await save_file_record(user["id"], file.filename, result["path"],
                                 file.content_type or "application/octet-stream",
                                 result.get("size", len(data)), kind)
    return {"file": rec, "url": f"/api/files/{rec['id']}"}


@api2.get("/files/{fid}")
async def download_file(fid: str, request: Request):
    # allow both cookie and ?auth= query for <img> tags
    rec = await db.files.find_one({"id": fid, "is_deleted": False}, {"_id": 0})
    if not rec: raise HTTPException(404, "File not found")
    try:
        data, ct = get_object(rec["storage_path"])
    except Exception as e:
        raise HTTPException(503, f"Storage error: {str(e)[:200]}")
    return Response(content=data, media_type=rec.get("content_type") or ct)


# --- AI Room Renders (Nano Banana) ---
SCENE_TEMPLATES = {
    "meeting_room": "modern corporate meeting room with executives seated at a conference table, sleek chairs, laptops on table, ambient lighting",
    "auditorium": "large modern auditorium with sloped tiered seating, audience seated, stage backdrop",
    "retail": "bright upscale retail store interior with mannequins, shoppers walking",
    "mall": "spacious indoor shopping mall atrium with escalators and hanging planters",
    "airport": "busy airport terminal with travelers and departure boards",
    "control_room": "high-tech mission control center with operators at consoles, dim ambient lighting",
    "school": "modern school lecture hall with students seated in rows",
    "hospital": "clean modern hospital reception lobby with waiting patients",
    "dooh": "urban rooftop with a large outdoor DOOH digital billboard against a night city skyline",
    "corporate": "premium corporate lobby with reception desk, marble floor and visitors",
    "broadcast_studio": "professional broadcast news studio with anchor desk and stage lighting",
}


class RenderInput(BaseModel):
    scene: str  # key from SCENE_TEMPLATES or custom prompt
    wall_width_m: float
    wall_height_m: float
    product_name: Optional[str] = None
    extra_prompt: Optional[str] = None


@api2.post("/ai/render")
async def ai_render(body: RenderInput, user=Depends(get_current_user)):
    import base64
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    scene_desc = SCENE_TEMPLATES.get(body.scene, body.scene)
    prompt = (
        f"Ultra-realistic photograph of {scene_desc}. Centered on the main wall is a large, "
        f"ultra-bright LED videowall (approximately {body.wall_width_m:.1f}m wide × "
        f"{body.wall_height_m:.1f}m tall) displaying a beautiful abstract "
        f"blue-and-purple corporate visual. The videowall is razor-thin bezel, edges perfectly "
        f"flush, showing brilliant HDR content. Professional architectural photography, "
        f"cinematic lighting, 8K, sharp focus, wide-angle. "
        f"{body.extra_prompt or ''}"
    )
    key = os.environ["EMERGENT_LLM_KEY"]
    chat = LlmChat(api_key=key, session_id=f"render-{user['id']}-{uuid.uuid4()}",
                   system_message="You generate photorealistic architectural renders.")
    chat.with_model("gemini", "gemini-3.1-flash-image-preview").with_params(modalities=["image", "text"])
    try:
        text, images = await chat.send_message_multimodal_response(UserMessage(text=prompt))
    except Exception as e:
        msg = str(e)
        if "Budget" in msg or "budget" in msg:
            raise HTTPException(402, "AI budget exceeded. Top up Emergent LLM key.")
        raise HTTPException(503, f"AI render failed: {msg[:200]}")
    if not images:
        raise HTTPException(500, "No image returned")
    img = images[0]
    data = base64.b64decode(img["data"])
    ct = img.get("mime_type", "image/png")
    ext = "png" if "png" in ct else "jpg"
    path = f"{APP_NAME}/renders/{user['id']}/{uuid.uuid4()}.{ext}"
    try:
        result = put_object(path, data, ct)
        rec = await save_file_record(user["id"], f"render-{body.scene}.{ext}", result["path"], ct, len(data), "render")
        return {"file_id": rec["id"], "url": f"/api/files/{rec['id']}", "scene": body.scene, "prompt": prompt}
    except Exception:
        # fallback: return base64 inline
        b64 = base64.b64encode(data).decode()
        return {"data_url": f"data:{ct};base64,{b64}", "scene": body.scene, "prompt": prompt}


# --- AI Content Generator ---
CONTENT_TYPES = {
    "tender_spec": "Draft a formal tender/RFP-style technical specification section for the LED videowall (bulleted, technical, government-tender tone).",
    "datasheet": "Draft a marketing-quality product datasheet body copy: hero paragraph + bulleted feature highlights + specifications summary.",
    "linkedin": "Draft a professional 3-paragraph LinkedIn post about a recent LED videowall project, hashtags at the end.",
    "proposal": "Draft a 5-paragraph executive project proposal / covering letter.",
    "brochure": "Draft a 2-page glossy brochure body copy: value proposition, applications, why LOGIC.",
    "email": "Draft a professional customer email pitching the videowall solution.",
    "comparison": "Draft a comparison sheet: our proposed LED wall vs a competitor's alternative — table + verdict.",
}


class ContentInput(BaseModel):
    content_type: str  # key from CONTENT_TYPES
    context: str       # project/product/site context provided by user
    tone: Optional[str] = "professional"


@api2.post("/ai/content")
async def ai_content(body: ContentInput, user=Depends(get_current_user)):
    from emergentintegrations.llm.chat import LlmChat, UserMessage
    kind = CONTENT_TYPES.get(body.content_type)
    if not kind: raise HTTPException(400, "Unknown content_type")
    sys = f"You are a technical + marketing writer for LOGIC LED Displays. {kind} Keep tone: {body.tone}."
    prompt = f"Context provided by user:\n{body.context}\n\nWrite the deliverable now."
    key = os.environ["EMERGENT_LLM_KEY"]
    chat = LlmChat(api_key=key, session_id=f"content-{user['id']}-{uuid.uuid4()}",
                   system_message=sys).with_model("anthropic", "claude-sonnet-4-6")
    try:
        text = await chat.send_message(UserMessage(text=prompt))
    except Exception as e:
        msg = str(e)
        if "Budget" in msg or "budget" in msg:
            raise HTTPException(402, "AI budget exceeded.")
        raise HTTPException(503, f"AI failed: {msg[:200]}")
    return {"content": text if isinstance(text, str) else str(text), "type": body.content_type}


@api2.get("/ai/content/types")
async def content_types(user=Depends(get_current_user)):
    return [{"key": k, "label": k.replace("_", " ").title(), "desc": v} for k, v in CONTENT_TYPES.items()]


@api2.get("/ai/render/scenes")
async def render_scenes(user=Depends(get_current_user)):
    return [{"key": k, "label": k.replace("_", " ").title()} for k in SCENE_TEMPLATES.keys()]



app.include_router(api2)
